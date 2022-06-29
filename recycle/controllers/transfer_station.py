from typing import Tuple

from django.conf import settings
from django.core.paginator import Paginator
from ninja import File, Query, Router
from ninja.errors import HttpError
from ninja.files import UploadedFile
from openpyxl import load_workbook
from openpyxl.cell import Cell

from infra.schemas import Page, Pagination
from recycle.models import Region, RegionGrade, RubbishVariety, StationNature, TransferStation
from recycle.schemas.transfer_station import TransferStationImportOut, TransferStationOut

router = Router(tags=["可回收物中转站"])


@router.get("", response=Pagination[TransferStationOut])
def list_stations(
    request,
    name: str = Query(None, title="中转站名称"),
    street_code: str = Query(None, title="街道编号"),
    community_code: str = Query(None, title="社区编号"),
    nature: StationNature = Query(None, title="场所性质"),
    page: Page = Query(...),
):
    """可回收物中转站台账列表"""

    stations = TransferStation.objects.order_by("-id")
    if name:
        stations = stations.filter(name__contains=name)
    if street_code:
        stations = stations.filter(street__code=street_code)
    if community_code:
        stations = stations.filter(community__code=community_code)
    if nature:
        stations = stations.filter(nature=nature)
    paginator = Paginator(stations, page.page_size)
    p = paginator.page(page.page)
    return {"count": paginator.count, "results": list(p.object_list)}


@router.post("/import", response={201: TransferStationImportOut})
def import_stations(request, file: UploadedFile = File(...)):
    """导入中转站台账"""

    columns = ["可回收中转站名称", "所属街道", "所属社区", "地址", "经度", "纬度", "运营单位", "负责人", "联系方式", "场所性质", "经营品种"]
    wb = load_workbook(filename=file)
    sheet = wb.worksheets[0]
    stations = list()

    column_row: Tuple[Cell] = sheet[1]
    if len(column_row) < len(columns):
        raise HttpError(400, f"录入错误，请检查表头是否符合\n {columns}")
    for idx, name in enumerate(columns):
        if column_row[idx].value != name:
            raise HttpError(400, f"录入错误，缺少'{name}'列")

    streets = Region.objects.filter(parent__code=settings.REGION_CODE, grade=RegionGrade.STREET)
    streets_dict = {r.name: r for r in streets}
    communities_dict = {r.name: r for r in Region.objects.filter(parent__in=streets)}

    content_row_start = 2
    for row in sheet.iter_rows(min_row=content_row_start):
        name = row[columns.index("可回收中转站名称")].value
        street_name = row[columns.index("所属街道")].value
        community_name = row[columns.index("所属社区")].value
        address = row[columns.index("地址")].value
        longitude = row[columns.index("经度")].value
        latitude = row[columns.index("纬度")].value
        management_company = row[columns.index("运营单位")].value
        manager_name = row[columns.index("负责人")].value
        manager_phone = row[columns.index("联系方式")].value
        nature = row[columns.index("场所性质")].value
        varieties = row[columns.index("经营品种")].value

        if street_name in streets_dict.keys():
            street = streets_dict[street_name]
        else:
            raise HttpError(400, f"录入错误，请检查街道名称：{street_name}。")

        if community_name in communities_dict.keys():
            community = communities_dict[community_name]
        else:
            raise HttpError(400, f"录入错误，请检查社区名称：{community_name}。")

        if nature not in StationNature.labels:
            raise HttpError(400, f"录入错误，场所性质仅支持 {StationNature.labels}")
        nature = [value for value, label in StationNature.choices if label == nature][0]

        varieties = set(varieties.split("、"))
        if varieties - set(RubbishVariety.labels):
            raise HttpError(400, f"录入错误，经营品种仅支持 {RubbishVariety.labels}")
        varieties = [RubbishVariety(value) for value, label in RubbishVariety.choices if label in varieties]

        station = TransferStation(
            name=name,
            street=street,
            community=community,
            address=address,
            longitude=longitude,
            latitude=latitude,
            management_company=management_company,
            manager_name=manager_name,
            manager_phone=manager_phone,
            nature=nature,
            varieties=varieties,
        )
        stations.append(station)
    TransferStation.objects.bulk_create(stations, ignore_conflicts=True)
    return TransferStationImportOut(imported_count=len(stations))
