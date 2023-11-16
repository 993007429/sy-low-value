import os

import csv
from typing import Tuple

from openpyxl import load_workbook
from openpyxl.cell import Cell
from django.conf import settings
from django.core.management import BaseCommand


class Command(BaseCommand):
    help = "创建出场记录文件(csv格式)"

    def add_arguments(self, parser):
        parser.add_argument(
            "-f",
            "--filepath",
            help="文件路径",
        )

        parser.add_argument(
            "-n",
            "--filename",
            help="新文件名",
        )

    def handle(self, filepath: str, filename: str, *args, **options):
        if not filepath:
            self.stdout.write("请提供文件路径")
            return

        if not filename:
            self.stdout.write("请提供新文件名")
            return

        if not filename.endswith(".csv"):
            self.stdout.write("新文件必须以.csv结尾")
            return

        columns = ["出场时间", "中转站名称", "司磅员", "车牌号",
                   "司机姓名", "运输单位", "可回收物类型", "细分品类",
                   "净重(kg)", "车牌识别照片", "车头照片","车顶照片", "去向"]
        labels = ['tare_weight_time','gross_weight_time','net_weight_time', 'station_id', 'weigher', 'plate_number', 'driver', 'carrier_name', 'recyclables_type'
            , 'category', 'tare_weight','gross_weight','net_weight', 'plate_number_photo_out', 'vehicle_head_photo_out', 'vehicle_roof_photo_out',
                  'place_to_go']
        csv_path = os.path.join(settings.BASE_DIR.parent, "recycle/migrations", filename)
        dct_arr = list()

        wb = load_workbook(filename=filepath)
        sheet = wb.worksheets[0]

        column_row: Tuple[Cell] = sheet[1][: len(columns)]
        column_row_values = [c.value for c in column_row]
        if column_row_values != columns:
            self.stdout.write(f"录入错误，请检查表头是否符合\n {columns}")
            return

        content_row_start = 2
        for row in sheet.iter_rows(min_row=content_row_start):
            dct = dict()
            dct["tare_weight_time"] = row[columns.index("出场时间")].value
            dct["gross_weight_time"] = row[columns.index("出场时间")].value
            dct["net_weight_time"] = row[columns.index("出场时间")].value
            dct["station_id"] = row[columns.index("中转站名称")].value
            dct["weigher"] = row[columns.index("司磅员")].value
            dct["plate_number"] = row[columns.index("车牌号")].value
            dct["driver"] = row[columns.index("司机姓名")].value
            dct["carrier_name"] = row[columns.index("运输单位")].value
            dct["recyclables_type"] = row[columns.index("可回收物类型")].value
            dct["category"] = row[columns.index("细分品类")].value
            dct["tare_weight"] = row[columns.index("净重(kg)")].value
            dct["gross_weight"] = row[columns.index("净重(kg)")].value
            dct["net_weight"] = row[columns.index("净重(kg)")].value
            dct["plate_number_photo_out"] = row[columns.index("车牌识别照片")].value
            dct["vehicle_head_photo_out"] = row[columns.index("车头照片")].value
            dct["vehicle_roof_photo_out"] = row[columns.index("车顶照片")].value
            dct["place_to_go"] = row[columns.index("去向")].value
            dct_arr.append(dct)

        with open(csv_path, "w") as file:
            writer = csv.DictWriter(file, fieldnames=labels)
            writer.writeheader()
            for elem in dct_arr:
                writer.writerow(elem)

        self.stdout.write(f"csv文件已生成: {csv_path}")
